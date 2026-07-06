"""Запись отчёта «Услуги-Деньги» (.xlsx) через openpyxl — заполнение шаблона.

Шаблон уже содержит формулы (D=G+I, F=C−J−H, G=F×E) и тарифы (E). Программа заполняет
только ВВОДНЫЕ ячейки листа «Услуги-Деньги» (C — всего из 071; H/I — частичники
кол-во/деньги; J — бесплатники кол-во) и листы Бесплатники/Частичники (клиенты + итоги).
"""

import os

import openpyxl
from openpyxl.utils import get_column_letter

from .model import CLIENT_COL, COL_TO_CANON, UD_ROW

CLIENT_COLS = sorted(COL_TO_CANON)   # колонки услуг на листах клиентов
ITOGO_COL = 15                        # «Итого услуг» (O)


def generate(template_path, out_path, data, period_text=None):
    template_path = os.path.abspath(template_path)
    out_path = os.path.abspath(out_path)
    wb = openpyxl.load_workbook(template_path)   # формулы шаблона сохраняются
    _fill_ud(wb["Услуги-Деньги"], data, period_text)
    _fill_free(wb["Бесплатники"], data.get("free_clients", []))
    _fill_partial(wb["Частичники"], data.get("part_cnt_clients", []),
                  data.get("part_money_clients", []))
    if os.path.exists(out_path):
        os.remove(out_path)
    wb.save(out_path)
    return out_path


def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _num_out(v):
    """Число для записи: целое — как int, дробное — округлённое до копеек; 0/пусто -> None."""
    if not v:
        return None
    v = float(v)
    return int(v) if v == int(v) else round(v, 2)


def _fill_ud(ws, data, period_text):
    if period_text:
        ws.cell(1, 2).value = period_text
    c071 = data.get("counts071", {})
    H = data.get("part_cnt_agg", {})
    I = data.get("part_money_agg", {})
    J = data.get("free_agg", {})
    prev = data.get("prev_ud", {})        # H/I/J предыдущего месяца (накопление)
    for key, row in UD_ROW.items():
        cc = ws.cell(row, 3)               # C — всего (из 071), не трогаем формулы
        if key in c071 and not _is_formula(cc):
            cc.value = c071[key]
        base = prev.get(key, {})
        # Накопительно: значение нового месяца = предыдущее + текущее.
        h = base.get("H", 0) + (H.get(key) or 0)   # кол-во частичная оплата
        i = base.get("I", 0) + (I.get(key) or 0)   # деньги частичная оплата
        j = base.get("J", 0) + (J.get(key) or 0)   # кол-во бесплатно
        if h:
            ws.cell(row, 8).value = _num_out(h)
        if i:
            ws.cell(row, 9).value = _num_out(i)
        if j:
            ws.cell(row, 10).value = _num_out(j)


def _client_row(ws, r, idx, fio, vals):
    ws.cell(r, 1).value = idx
    ws.cell(r, 2).value = fio
    for col in CLIENT_COLS:
        v = vals.get(COL_TO_CANON[col], 0)
        ws.cell(r, col).value = (round(v, 2) if isinstance(v, float) and v != int(v)
                                 else int(v)) if v else None
    ws.cell(r, ITOGO_COL).value = f"=SUM(C{r}:N{r})"


def _totals_row(ws, r, first, last):
    for col in CLIENT_COLS + [ITOGO_COL]:
        L = get_column_letter(col)
        ws.cell(r, col).value = f"=SUM({L}{first}:{L}{last})"


def _unmerge_all(ws):
    """Разъединить объединённые ячейки (в них нельзя писать; текст остаётся в левой-верхней)."""
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))


def _fill_free(ws, clients):
    _unmerge_all(ws)
    r = 4
    for i, (fio, vals) in enumerate(clients, 1):
        _client_row(ws, r, i, fio, vals)
        r += 1


def _fill_partial(ws, cnt_clients, money_clients):
    _unmerge_all(ws)
    # заголовок блока (r1..r3) — для повторного использования под блоком денег
    hdr = [[ws.cell(rr, cc).value for cc in range(1, 16)] for rr in (1, 2, 3)]
    r = 4
    first = r
    for i, (fio, vals) in enumerate(cnt_clients, 1):
        _client_row(ws, r, i, fio, vals)
        r += 1
    if cnt_clients:
        _totals_row(ws, r, first, r - 1)
        r += 1
    r += 1  # пустая строка-разделитель
    for hrow in hdr:                       # заголовок блока денег
        for cc, val in enumerate(hrow, 1):
            if val is not None:
                ws.cell(r, cc).value = val
        r += 1
    first = r
    for i, (fio, vals) in enumerate(money_clients, 1):
        _client_row(ws, r, i, fio, vals)
        r += 1
    if money_clients:
        _totals_row(ws, r, first, r - 1)
