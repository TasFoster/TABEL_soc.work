"""Чтение входных файлов функции «Услуги-Деньги».

* read_071(.xls)  — сводка количества услуг -> {canon: количество} (xlrd).
* read_benefit(.xlsx) — листы «бесплатники»/«частичники» -> клиенты и агрегаты по
  услугам (openpyxl). У частичников два блока: количества и деньги.
"""

import openpyxl
import xlrd

from .model import canon_of


# ----------------------------------------------------------------- 071 (.xls)
def read_071(path):
    """Вернуть {canon: количество услуг} из отчёта 071."""
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    hdr, best = None, 0
    for r in range(sh.nrows):
        n = sum(1 for c in range(sh.ncols) if canon_of(sh.cell_value(r, c)))
        if n > best:
            best, hdr = n, r
    counts = {}
    if hdr is None:
        return counts
    for c in range(sh.ncols):
        k = canon_of(sh.cell_value(hdr, c))
        if not k:
            continue
        v = sh.cell_value(hdr + 1, c) if hdr + 1 < sh.nrows else 0
        try:
            counts[k] = counts.get(k, 0) + float(v or 0)
        except (TypeError, ValueError):
            pass
    return counts


# ----------------------------------------- единый загрузчик книги (.xlsx и .xls)
# read_benefit писался под openpyxl (1-based ws.cell(r,c).value, ws.max_row/max_column,
# ws.title, wb.worksheets). Чтобы принимать и .xls, оборачиваем xlrd в тот же интерфейс.
class _XlsCell:
    __slots__ = ("value",)

    def __init__(self, value):
        self.value = value


class _XlsSheet:
    def __init__(self, sheet):
        self._s = sheet
        self.title = sheet.name
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def cell(self, row, column):
        r, c = row - 1, column - 1            # openpyxl 1-based -> xlrd 0-based
        if 0 <= r < self._s.nrows and 0 <= c < self._s.ncols:
            v = self._s.cell_value(r, c)
            return _XlsCell(None if v == "" else v)   # пустую строку -> None (как openpyxl)
        return _XlsCell(None)


class _XlsBook:
    def __init__(self, path):
        wb = xlrd.open_workbook(path)
        self.worksheets = [_XlsSheet(wb.sheet_by_index(i)) for i in range(wb.nsheets)]


def _load_book(path):
    """Книга с openpyxl-совместимым интерфейсом: .xls -> xlrd-адаптер, иначе openpyxl."""
    if (path or "").lower().endswith(".xls"):
        return _XlsBook(path)
    return openpyxl.load_workbook(path, data_only=True)


# --------------------------------------------------------------- льготники (.xlsx/.xls)
def _sheet(wb, *names):
    for nm in names:
        for ws in wb.worksheets:
            if nm.lower() in (ws.title or "").lower():
                return ws
    return None


def _colmap(ws, hr):
    """Колонка -> canon по строке-заголовку hr."""
    m = {}
    for c in range(1, ws.max_column + 1):
        k = canon_of(ws.cell(hr, c).value)
        if k and k not in m.values():
            m[c] = k
    return m


def _header_rows(ws):
    """Строки-заголовки колонок услуг (>=3 совпадений canon)."""
    out = []
    for r in range(1, ws.max_row + 1):
        n = sum(1 for c in range(1, ws.max_column + 1) if canon_of(ws.cell(r, c).value))
        if n >= 3:
            out.append(r)
    return out


def _parse_block(ws, hr):
    """Клиенты блока, начиная со строки заголовка hr. -> ([(fio, {canon: val})], colmap)."""
    cm = _colmap(ws, hr)
    clients = []
    for r in range(hr + 1, ws.max_row + 1):
        fio = ws.cell(r, 2).value
        n1 = ws.cell(r, 1).value
        is_client = fio and str(fio).strip() and isinstance(n1, (int, float))
        if is_client:
            vals = {}
            for c, k in cm.items():
                v = ws.cell(r, c).value
                vals[k] = float(v) if isinstance(v, (int, float)) else 0.0
            clients.append((str(fio).strip(), vals))
        elif clients:
            break  # итоги/следующий блок/пусто после клиентов — конец блока
    return clients, cm


def _aggregate(clients):
    agg = {}
    for _fio, vals in clients:
        for k, v in vals.items():
            agg[k] = agg.get(k, 0) + (v or 0)
    return agg


def read_benefit(path):
    """Вернуть данные по бесплатникам и частичникам (клиенты + агрегаты по услугам).

    Принимает и .xlsx (openpyxl), и .xls (xlrd-адаптер) — структура листов одинакова."""
    wb = _load_book(path)
    out = {"free_clients": [], "free_agg": {},
           "part_cnt_clients": [], "part_cnt_agg": {},
           "part_money_clients": [], "part_money_agg": {}}
    free = _sheet(wb, "бесплатник")
    if free is not None:
        hrs = _header_rows(free)
        if hrs:
            out["free_clients"], _ = _parse_block(free, hrs[0])
            out["free_agg"] = _aggregate(out["free_clients"])
    part = _sheet(wb, "частичник")
    if part is not None:
        hrs = _header_rows(part)
        if hrs:  # первый блок — количества
            out["part_cnt_clients"], _ = _parse_block(part, hrs[0])
            out["part_cnt_agg"] = _aggregate(out["part_cnt_clients"])
        if len(hrs) >= 2:  # второй блок — деньги
            out["part_money_clients"], _ = _parse_block(part, hrs[1])
            out["part_money_agg"] = _aggregate(out["part_money_clients"])
    return out


# ----------------------------------------------- предыдущий отчёт (для накопления)
def _num(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def read_prev_ud(path):
    """H/I/J предыдущего отчёта «Услуги-Деньги» по услугам -> {canon: {'H':..,'I':..,'J':..}}.

    Нужно для НАКОПИТЕЛЬНОГО отчёта: H/I/J нового месяца = значения предыдущего + новые.
    H — кол-во частичная (кол.8), I — деньги частичная (кол.9), J — кол-во бесплатно (кол.10);
    услуга опознаётся по названию (кол.2). Поддерживает .xlsx (этой программы) и .xls."""
    low = (path or "").lower()
    rows = _prev_rows_xlsx(path) if low.endswith(".xlsx") else _prev_rows_xls(path)
    out = {}
    for name, h, i, j in rows:
        key = canon_of(name)
        if key and key not in out:
            out[key] = {"H": _num(h), "I": _num(i), "J": _num(j)}
    return out


def _prev_rows_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next((w for w in wb.worksheets if "деньги" in (w.title or "").lower()),
              wb.worksheets[0])
    return [(ws.cell(r, 2).value, ws.cell(r, 8).value, ws.cell(r, 9).value, ws.cell(r, 10).value)
            for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value]


def _prev_rows_xls(path):
    wb = xlrd.open_workbook(path)
    sh = None
    for i in range(wb.nsheets):
        if "деньги" in (wb.sheet_by_index(i).name or "").lower():
            sh = wb.sheet_by_index(i)
            break
    sh = sh or wb.sheet_by_index(0)
    out = []
    for r in range(sh.nrows):
        name = sh.cell_value(r, 1) if sh.ncols > 1 else None
        if name:
            h = sh.cell_value(r, 7) if sh.ncols > 7 else 0
            i = sh.cell_value(r, 8) if sh.ncols > 8 else 0
            j = sh.cell_value(r, 9) if sh.ncols > 9 else 0
            out.append((name, h, i, j))
    return out
