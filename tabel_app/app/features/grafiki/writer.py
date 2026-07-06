"""Запись «Графика проверок» (.xlsx) с нуля через openpyxl."""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

MONTHS_NOM = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
              "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]


def generate(out_path, ctx, workers, weeks, marks):
    out_path = os.path.abspath(out_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "График проверок"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    ncols = 2 + len(weeks)
    last = get_column_letter(ncols)

    # ---- шапка «Утверждаю» (справа)
    r = 1
    for line in ("«Утверждаю»", "Директор филиала", ctx["org_full"], ctx["district"],
                 f"________ {ctx['director']}"):
        ws.cell(r, 5, line)
        r += 1
    r += 1

    # ---- заголовок
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    title = (f"График проверок качества работы социальных работников отделения № "
             f"{ctx['dept_no']} на {ctx['half']} полугодие {ctx['year']} года")
    tc = ws.cell(r, 1, title)
    tc.alignment = center
    tc.font = bold
    r += 2

    head_month = r
    head_date = r + 1
    first_data = r + 2

    # № и ФИО — объединить по двум строкам шапки
    ws.merge_cells(start_row=head_month, start_column=1, end_row=head_date, end_column=1)
    ws.merge_cells(start_row=head_month, start_column=2, end_row=head_date, end_column=2)
    ws.cell(head_month, 1, "№").alignment = center
    ws.cell(head_month, 2, "ФИО социального работника").alignment = center

    # месяцы (объединение по неделям месяца) + даты
    col = 3
    i = 0
    while i < len(weeks):
        m = weeks[i][0]
        j = i
        while j < len(weeks) and weeks[j][0] == m:
            j += 1
        span = j - i
        ws.merge_cells(start_row=head_month, start_column=col,
                       end_row=head_month, end_column=col + span - 1)
        mc = ws.cell(head_month, col, MONTHS_NOM[m])
        mc.alignment = center
        mc.font = bold
        for k in range(span):
            ws.cell(head_date, col + k, weeks[i + k][1]).alignment = center
        col += span
        i = j

    # ---- строки соцработников
    rr = first_data
    for idx, w in enumerate(workers, 1):
        ws.cell(rr, 1, idx).alignment = center
        ws.cell(rr, 2, w["fio"])
        if w.get("self_control"):
            cc = ws.cell(rr, 3, "самоконтроль")
            ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=ncols)
            cc.alignment = center
        else:
            for wk in range(len(weeks)):
                v = marks.get((idx - 1, wk))
                if v:
                    ws.cell(rr, 3 + wk, v).alignment = center
        rr += 1
    last_data = rr - 1

    # ---- границы по сетке (шапка + данные)
    for row in range(head_month, last_data + 1):
        for c in range(1, ncols + 1):
            ws.cell(row, c).border = border

    # ---- подпись
    rr += 1
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=min(ncols, 12))
    ws.cell(rr, 1, f"Заведующая отделением социального обслуживания на дому № "
                   f"{ctx['dept_no']}        ________  {ctx['zav']}")

    # ---- ширины/печать
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    for c in range(3, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 4.6
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    if os.path.exists(out_path):
        os.remove(out_path)
    wb.save(out_path)
    return out_path
